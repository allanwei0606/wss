#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# @Time    : 7/22/2022 2:26 PM
# @Author  : Allan Wei
# @File    : testData.py

from openpyxl import Workbook
from faker import Faker
import random
import time

#写excel操作
def write_excel(fileName,data):

    wb = Workbook()
    # 写入表头
    excel_head = ['note', 'phone', 'fullname', 'email', 'tag','School','Skills','Degree','Major','Title']
    sheet0Name = 'nameList'

    sheet0 = wb.create_sheet(sheet0Name, index=0)
    for i, item in enumerate(excel_head):
        sheet0.cell(row=1, column=i + 1, value=item)

    # 写入数据
    for i, item in enumerate(data):
        i = i + 2
        for j, val in enumerate(item):
            sheet0.cell(row=i, column=j + 1, value=val)

    wb.save(fileName + '.xlsx')


#生成比例手机号
def maker_phone():
    num_113 = ["113" for i in range(3)]
    num_112 = ["112" for j in range(3)]
    num_111 = ["111" for k in range(4)]
    phone_num_head = num_113 + num_112 + num_111
    phone_head = random.sample(phone_num_head,1)
    telephone = phone_head[0]
    phone_tail = []
    for i in range(8):
        slice = random.randint(0,9)
        phone_tail.append(str(slice))

    res = telephone + "".join(phone_tail)
    return res

#生成测试数据
def faker_maker(num,i):
    f = Faker()

    print(num,i)
    faker_list = []
    phone1 = maker_phone()
    name = f.name()
    num_three = f.numerify()
    email =f.email()



    faker_list.append(str(i+1))
    faker_list.append(phone1)
    faker_list.append(name)
    faker_list.append(email)
    faker_list.append(num_three)
    return faker_list

# 将测试数据放入列表
def faker_list_add(num):
    base_list = []
    for i in range(num):
        fake_new_list = faker_maker(num,i)
        base_list.append(fake_new_list)
    return base_list

start_time = time.time()
write_excel(r'D:\workspace\ool1',faker_list_add(50000))
end_time = time.time()
total_time = end_time - start_time
print(total_time)
